import os
import jieba
import re
import csv
import openpyxl
from openpyxl.styles import Border, Side, colors
from openpyxl import load_workbook
from openpyxl.cell import Cell
global word_meaning_map
global synonym_word_map
word_meaning_map = dict()
synonym_word_map = dict()

CN_NUM = {
    '1': '一', '2': '二', '3': '三', '4': '四', '5': '五', '6': '六', '7': '七', '8': '八', '9': '九', '0': '零',
}

def handle_digital(word):
    '''
    数字规整
    目前可以转小数、百分数、正数，最高只能转到正数三位数，如999%，999，999.0
    :param word:待转的文字
    :return:
    '''
    pattern1 = re.compile('\d.\d')  #对小数进行处理
    pattern2 = re.compile('\d%')  #对百分之的数据进行处理

    # print(word)
    # print(type(word))
    if word != None:
        if "廿" in word:
            word = word.replace("廿","二十")
        if re.search(pattern2, word):
            str_list = list(word)
            str_list.remove('%')
            new_str_list = list()
            n = 0
            for i in str_list:  # 将百分号放在数字前面
                if i.isdigit() and n == 0:
                    new_str_list.append("%")
                    new_str_list.append(i)
                    n = n + 1
                else:
                    new_str_list.append(i)
            word = ''.join(new_str_list)
            word = word.replace('%', '百分之')
        if re.search(pattern1, word):#将小数点转为汉字”点“
            #word = word.replace(".", "点")
            if "." in word:
                word_zhengshu = handle_zhengshu(word.split(".")[0])
                word_xiaoshu = word.split(".")[-1]
                #print(f"word_xiaoshu:{word_xiaoshu}")
                shuzi_list = list()
                for xiaoshu in word_xiaoshu:
                    if xiaoshu.isdigit():
                        shuzi = CN_NUM[xiaoshu]
                        shuzi_list.append(shuzi)
                    else:
                        shuzi_list.append(xiaoshu)
                #print(f"shuzi_list:{shuzi_list}")
                word_xiaoshu = "".join(shuzi_list)
                word = word_zhengshu+"点"+word_xiaoshu
                #print(f"word1:{word}")
        word = handle_zhengshu(word)

    return word

def handle_zhengshu(word):
    #正数1-3位进行转换
    pattern3 = re.compile('\d*')  # 对数字进行处理
    if re.search(pattern3, word):
        str_list = list(word)
        n = 0
        str_index = 0
        new_str_list = list()
        digit_to_str = str()
        digit_str = str()
        for i in range(0, len(str_list)):  # 将字符串内的数字扣出来
            if str_list[i].isdigit():
                #print(f"{word}出现数字:{str_list[i]}")
                if n == 0:
                    str_index = i  # 从第一个出现的数字开始记录index
                    new_str_list.append(str_list[i])  # 第一个出现的数字存到new_str_list
                    digit_str = str_list[i]  # 第一个出现的数字赋值给digit_str
                if i > 0 and str_list[i - 1] == digit_str:  # 之后出现的数字判断下跟上一个数字是否是连续出现的
                    digit_str = str_list[i]  # 是连续出现的再赋值给digit_str
                    new_str_list.append(digit_str)
                n = n + 1
        #print(f"new_str_list{new_str_list}")

        digit_to_str_list = list()
        if new_str_list!= [] and str(new_str_list[0]) == str(0):  # 若第一个数字是0，则直接按单个数字转
            for get_shuzi in new_str_list:
                shuzi = CN_NUM[get_shuzi]
                digit_to_str_list.append(shuzi)
            print(digit_to_str_list)
            digit_to_str = "".join(digit_to_str_list)

        elif len(new_str_list) == 2:  # 对两位数进行转换
            n = 0
            for i in range(0, len(new_str_list)):
                if i == 0 and int(new_str_list[i]) == 1:  # 两位数第一位1不转换
                    pass
                elif i == 1 and int(new_str_list[i]) == 0:  # 两位数第二位0不转换
                    pass
                else:
                    digit_to_str_list.append(CN_NUM[new_str_list[i]])
                if n == 0:  # 从第一位数字后加”十“
                    digit_to_str_list.append("十")
                n = n + 1
            digit_to_str = ''.join(digit_to_str_list)
            # print(digit_to_str)
        elif len(new_str_list) == 1:  # 对一位数进行转换
            digit_to_str_list.append(CN_NUM[new_str_list[0]])
            digit_to_str = CN_NUM[new_str_list[0]]

        elif len(new_str_list) == 3:  # 对三位数进行转换
            if int(new_str_list[0]) == 2:  # 三位数第一位2转两
                digit_to_str_list.append("两百")
            else:
                digit_to_str_list.append(CN_NUM[new_str_list[0]] + "百")
            if int(new_str_list[1]) == 0 and int(new_str_list[2]) == 0:
                pass
            elif int(new_str_list[1]) == 0:
                digit_to_str_list.append("零")
            elif int(new_str_list[1]) != 0:
                digit_to_str_list.append(CN_NUM[new_str_list[1]] + "十")
            if int(new_str_list[2]) == 0:
                pass
            else:
                digit_to_str_list.append(CN_NUM[new_str_list[2]])
            digit_to_str = ''.join(digit_to_str_list)
            # print(f"三位数:{digit_to_str}")

        if digit_to_str_list != []:
            change_list = list()
            for i in range(0, len(str_list)):
                if str_index == i and len(new_str_list) > 0:
                    change_list.append(digit_to_str)
                elif len(new_str_list) == 2 and i == str_index + 1:
                    pass
                elif len(new_str_list) == 3 and i == str_index + 1:
                    pass
                elif len(new_str_list) == 3 and i == str_index + 2:
                    pass
                else:
                    change_list.append(str_list[i])
            word = ''.join(change_list)
    return word

def word_meaming_switch_to_word(word):
    '''
    语义相近词返回表头
    :param word:
    :return:
    '''
    for key in word_meaning_map:
        #print(key)
        meaning_list = word_meaning_map[key]
        #print(meaning_list)
        if word in meaning_list:
            word = key
            return word
    word = synonym_word_switch(word)
    return word

def word_meaming_switch(expect: str, actual: str) -> bool:
    '''
    判断是否语义正确
    :param expect:
    :param actual:
    :return:
    '''
    for key in word_meaning_map:
        #print(key)
        meaning_list = word_meaning_map[key]
        #print(meaning_list)
        if expect == actual:
            return True
        if expect in meaning_list:
            #print(meaning_list)
            if actual in meaning_list:
                #print(meaning_list)
                return True
        expect = synonym_word_switch(expect)
        actual = synonym_word_switch(actual)
        if expect == actual:
            return True
    return False

#查找替换同义词
def synonym_word_switch(word):
    #先进行数字规整
    if word != None:
        word = handle_digital(word)
        fenci_lsit = []
        for key in synonym_word_map:
            synonym_word_list = synonym_word_map[key]
            #将列表内所有值转为字符串格式
            synonym_word_list_1 = [str(i) for i in synonym_word_list]
            fenci_lsit = []
            pipei_count = 0
            fenci = jieba.lcut(word, cut_all=False)
            for i in fenci:
                if i in synonym_word_list_1:
                    i = key
                    pipei_count = pipei_count + 1
                    fenci_lsit.append(i)
                else:
                    fenci_lsit.append(i)
            if pipei_count > 0:
                word = "".join('%s' %i for i in fenci_lsit)
                continue
        if len(fenci_lsit) != 0:
            word = "".join('%s' %i for i in fenci_lsit)
            # 去掉字符串内所有的空格
            word = "".join(re.split(r'\s+', word))
        else:
            # 去掉字符串内所有的空格
            word = "".join(re.split(r'\s+',word))
    return word

def load_map(filename,dictpath):
    if os.path.exists(dictpath):
        jieba.load_userdict(dictpath)
    else:
        print(f"{dictpath}路径不存在！")
    wb = load_workbook(filename)
    first_sheet = wb.sheetnames[0]
    ws = wb[first_sheet]
    global word_meaning_map
    last_word = ''
    for row_index, row_item in enumerate(ws.rows):
        if row_index == 0 or len(row_item) != 2:
            continue
        first_item: Cell = row_item[0]
        if first_item.value:
            last_word = first_item.value
            #print("last_word=", last_word)
            word_meaning_map[last_word] = list()
            word_meaning_map[last_word].append(last_word)
        if row_item[1].value:
            if row_item[1].value not in word_meaning_map[last_word]:
                word_meaning_map[last_word].append(row_item[1].value)
    # print(word_meaning_map)

    second_sheet = wb.sheetnames[1]
    ws_second = wb[second_sheet]
    global synonym_word_map
    for row_index, row_item in enumerate(ws_second.rows):
        if row_index == 0 or len(row_item) != 2:
            continue
        first_item: Cell = row_item[0]
        if first_item.value:
            synonym_word = first_item.value
            #print("synonym_word=", synonym_word)
            synonym_word_map[synonym_word] = list()
            synonym_word_map[synonym_word].append(synonym_word)
        if row_item[1].value:
            if row_item[1].value not in synonym_word_map[synonym_word]:
                synonym_word_map[synonym_word].append(row_item[1].value)
    # print(synonym_word_map)


# if __name__ == "__main__":
#     jieba.set_dictionary("../config/dict_all.txt")
#     jieba.initialize()
#     load_map('../config/语义表.xlsx', '../config/new_dict.txt')
#     print(word_meaming_switch_to_word("hello world你今天怎么样"))
#     #print(synonym_word_switch("16度"))
#     #print(word_meaming_switch("台灯打开照明电脑模式","打开照明电脑模式"))
#     # file_list = os.listdir("./data")
#     # print(file_list)
#     # for getfile in file_list:
#     #     if ".csv" in getfile:
#     #         filepath = os.path.join("./data",getfile)
#     #         print(filepath)
#     #         print(os.path.split(filepath))
#     #         run(filepath)




