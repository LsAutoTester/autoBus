# _*_ coding: utf-8 _*_
# @Time : 2022/5/16 13:34
# @Author : xjzhu
# @File : handler_excel.py
# @desc : excel文件处理：使用xlrd读取数据，openpyxl写数据
import os,re
import shutil
import time

import xlrd
import openpyxl
from openpyxl.styles import Border, Side, colors,Alignment
#excel文件处理

#如内容含音标打印，需用以下代码
# sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='gb18030')
#获取当前时间
DATE_TIME = time.strftime('%y%m%d_%H%M%S')

def is_number(string):
    pattern = re.compile(r'^[-+]?[0-9]*\.?[0-9]+([eE][-+]?[0-9]+)?$')
    return bool(pattern.match(string))


class Handler_excel():
    def __init__(self,file):
        """

        @param filename: 文件的全路径
        @param sheetname: 表名
        """
        # FileDirectory = os.path.dirname(os.path.abspath(__file__))  # 获取当前脚本所在目录的绝对路径（文件目录）
        # self.data_filedir = os.path.dirname(FileDirectory) + '\\test_data\\'  # r 不转义
        # self.filename = self.data_filedir + filename
        # self.sheetname = sheetname
        self.file = file
        self.fill_red = openpyxl.styles.PatternFill('solid', fgColor="dc143c")
        self.fill_green = openpyxl.styles.PatternFill('solid', fgColor="7cfc00")
        self.fill_blue = openpyxl.styles.PatternFill('solid', fgColor="0000ff")
        self.LightSkyBlue = openpyxl.styles.PatternFill('solid', fgColor="87CEFA")
        self.Border_1 = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'))
        self.Alignment_center = Alignment(horizontal='center', vertical='center')
        if not os.path.isfile(self.file):
            self.creat_resultfile(self.file)
        self.wordbook = openpyxl.load_workbook(self.file)  # 如果从xls转为xlsx，会出现warning，可忽略。
        self.sheetname = self.wordbook.sheetnames[0]

    # 将text转为url链接
    def excel_text2url(self,link_url):
        ''' 把一个网址字符串转换为excel公式 '''
        return f'=HYPERLINK("{link_url}")'

    def check_file(self,*args):
        """
        检查文件是否存在
        @param args: 0:文件路径；1：表格名
        @return:
        """
        if (args):
            filename =  args[0]
            sheetname = args[1]
        else:
            filename = self.file
            sheetname = self.sheetname
        try:
            workbook = openpyxl.load_workbook(filename=filename)
            if (sheetname not in workbook.sheetnames):
                print("当前测试文件:{0}\n未找到表名：{1}".format(filename,sheetname))
                return False
            else:
                print("当前测试文件名：{0}\n，表名：{1}".format(filename, sheetname))
            return True
        except Exception as e:
            if (os.path.exists(filename) == False):
                print("文件路径不存在，请检查：", filename)
            print("check_file发生异常信息：",e)
            return False

    #使用xlrd读取数据，速度更优
    def read_data_byxlrd(self):
        """
        使用xlrd读取excel数据，把内容转为字典格式的数据
        @return: 返回测试用例列表（字典类型组成的列表）
        """
        book = openpyxl.load_workbook(self.file)
        # print(f'表名：{book.sheet_names()}')
        sheetnames = book.sheetnames
        if self.sheetname not in sheetnames:
            print(f"找不到'{self.sheetname}'这个sheet页")
            return False
        sheet = book[self.sheetname]
        # sheetname = sheet.name      #获取当前表格的表名
        # 获取第一行表头
        # print("表头标题：",title)
        nrows = sheet.max_row + 1     #获取该表的总行数
        ncols = sheet.max_column + 1     #获取该表的总列数
        title = []
        for r in range(1,ncols):
            title.append(sheet.cell(1,r).value)
        #读取表头之外的数据
        cases = []
        #遍历第一行之外的其他行
        for i in range(2,nrows):
            row_value_list = []
            for r in range(1,ncols):
                row_value = sheet.cell(i,r).value
                if row_value == '':
                    row_value = sheet.hyperlink_map.get((i, r))
                if row_value == None:
                    row_value = ''
                row_value_list.append(row_value)    #获取该行中所有的单元格对象组成的列表
            dic = dict(zip(title,row_value_list))     #通过zip聚合，dict函数将列表转换为字典
            cases.append(dic)       #将获取的字典添加到字典中
        return cases

    def merge_cells(self,sr,sc,er,ec):
        sh = self.wordbook[self.sheetname]
        sh.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)

        # 复制文件
    def save_file(self):
        self.wordbook.save(self.file)

    def write_data(self,*args):
        """
        openpyxl注意事项：表格第0行为文档密级，第1行才是表头。列也是从1开始。
        @param args: 0：指定行；1：指定列;2:需写入的数据
        @return: 无
        """
        #操作文件进行写数据
        try:
            # wordbook = openpyxl.load_workbook(self.file)     #如果从xls转为xlsx，会出现warning，可忽略。
            sh = self.wordbook[self.sheetname]
            # 确定要操作的单元格,并给定写入的值
            sh.cell(row=args[0],column=args[1],value = args[2])
            sh.cell(row=args[0], column=args[1]).alignment = self.Alignment_center
            if args[2] == "Fail" or args[2] == "FAIL":
                sh.cell(row=args[0],column=args[1]).fill = self.fill_red
            if args[2] == "Pass" or args[2] == "PASS":
                sh.cell(row=args[0], column=args[1]).fill = self.fill_green

            #将缓冲区中的数据立刻写入文件，同时清空缓冲区
            #sh.flush()
            #保存内容
            #self.wordbookwordbook.save(self.file)
            # print('内容已保存：{}。'.format(value))
        except Exception as e:
            print(f'write_data()报错异常信息：{e}')

    def Number_format(self,*args):
        try:
            sh = self.wordbook[self.sheetname]
            sh.cell(row=args[0], column=args[1]).number_format = '0.00%'
        except Exception as e:
            print(f'Number_format()报错异常信息：{e}')

    def style_format(self,*args):
        try:
            sh = self.wordbook[self.sheetname]
            sh.cell(row=args[0], column=args[1]).border = self.Border_1
            sh.cell(row=args[0], column=args[1]).fill = self.LightSkyBlue
        except Exception as e:
            print(f'Number_format()报错异常信息：{e}')

    def append_data(self,data):
        try:
            sh = self.wordbook[self.sheetname]
            sh.append(data)
        except Exception as e:
            print(f'append_data()报错异常信息：{e}')

    def get_rowclo(self):
        book = xlrd.open_workbook(self.file)
        sheet = book.sheet_by_name(self.sheetname)
        nrows = sheet.nrows
        nclos = sheet.ncols
        return nrows,nclos

    def copy_file(self,oldfile, newfile):
        """
        复制文件，生成一个新的文件（可用于保存结果）。
        @param oldfile: 原文件全路径
        @param newfile:   新文件路径
        @return:   newfile 新文件路径
        """
        if (os.path.exists(newfile)):  # 如果存在文件，先删除
            os.remove(newfile)
        shutil.copy(oldfile, newfile)
        # print('已复制为新的文件：{}'.format(new_filedir))
        # return newfile


    def creat_resultfile(self,RESULT_File):
        wb = openpyxl.Workbook()
        wb.create_sheet(self.sheetname )
        wb.save(RESULT_File)
        sheet = wb["Sheet"]
        wb.remove(sheet)
        wb.save(RESULT_File)
        #self.copy_file(self.file, new_file)


# if __name__ == "__main__":
#     date = time.strftime('%m%d_%H%M_')
#     # casefile = os.path.join(DATA_DIR, '用例.xlsx')
#     casefile = os.path.join(DATA_DIR, '单词卡固件测试用例.xlsx')
#     sheetname = '单词音标表'
#     excel = Handler_excel(casefile, sheetname)
#     cases = excel.read_data_byxlrd()  # 测试用例列表
#     file = excel.creat_resultfile()
#     print(file)
#     # excel.copy_file(oldfile, newfile):


