import openpyxl
from openpyxl.styles import Border, Side, colors
import ctypes  # 改变输出文字的颜色
import sys


# 设置excel表格样式
class excel_styel():
    def __init__(self):
        self.border_set_bottom = Border(left=Side(style='medium', color=colors.WHITE),
                              right=Side(style='medium', color=colors.BLACK),
                              top=Side(style='medium', color=colors.WHITE),
                              bottom=Side(style='medium', color=colors.BLACK))

        self.fill_test = openpyxl.styles.PatternFill('solid', fgColor="0000ff")  # 颜色也可以直接设置red等
        self.fill_red = openpyxl.styles.PatternFill('solid', fgColor="dc143c")
        self.fill_green = openpyxl.styles.PatternFill('solid', fgColor="7cfc00")
        self.fill_blue = openpyxl.styles.PatternFill('solid', fgColor="0000ff")
        self.fill_skybule = openpyxl.styles.PatternFill('solid', fgColor="87ceeb")



STD_INPUT_HANDLE = -10
STD_OUTPUT_HANDLE = -11
STD_ERROR_HANDLE = -12
# Windows CMD命令行 字体颜色定义 text colors
FOREGROUND_BLACK = 0x00  # black.
FOREGROUND_BLUE = 0x09  # blue.
FOREGROUND_GREEN = 0x0a  # green.
FOREGROUND_RED = 0x0c  # red.
FOREGROUND_YELLOW = 0x0e  # yellow.
FOREGROUND_WHITE = 0x0f  # white.

# Windows CMD命令行 背景颜色定义 background colors
BACKGROUND_YELLOW = 0xe0  # yellow.
BACKGROUND_BLUE = 0x10  #blue.
BACKGROUND_GREEN = 0x20  #green.
BACKGROUND_RED = 0x40  #red.
BACKGROUND_INTENSITY = 0x80 #intensity
BACKGROUND_WHITE = 0xF0  #white.

std_out_handle = ctypes.windll.kernel32.GetStdHandle(STD_OUTPUT_HANDLE)


def set_cmd_text_color(color, handle=std_out_handle):
    Bool = ctypes.windll.kernel32.SetConsoleTextAttribute(handle, color)
    return Bool

def resetColor():
    # reset white
    set_cmd_text_color(FOREGROUND_RED | FOREGROUND_GREEN | FOREGROUND_BLUE)


class cmd_output():
    def printGreen(self,mess):
        '''
        # 绿色
        :param mess:
        :return:
        '''
        set_cmd_text_color(FOREGROUND_GREEN)
        sys.stdout.write(mess)
        resetColor()


    def printRed(self,mess):
        '''
        # 红色
        :param mess:
        :return:
        '''
        set_cmd_text_color(BACKGROUND_RED)
        sys.stdout.write(mess)
        resetColor()


    def printYellowRed(self,mess):
        '''
        # 黄底蓝字
        :param mess:
        :return:
        '''
        set_cmd_text_color(BACKGROUND_YELLOW | FOREGROUND_RED)
        sys.stdout.write(mess)
        resetColor()

    def printYellow(self,mess):
        '''
        # 黄字
        :param mess:
        :return:
        '''
        set_cmd_text_color(FOREGROUND_YELLOW)
        sys.stdout.write(mess)
        resetColor()


    def printBlue(self,mess):
        '''
        # 蓝色
        :param mess:
        :return:
        '''
        set_cmd_text_color(FOREGROUND_BLUE)
        sys.stdout.write(mess)
        resetColor()

    def printIntensity(self,mess):
        '''
        # 高亮背景
        :param mess:
        :return:
        '''
        set_cmd_text_color(BACKGROUND_INTENSITY)
        sys.stdout.write(mess)
        resetColor()

    def printWhiteBlue(self,mess):
        '''
        # 白底蓝字
        :param mess:
        :return:
        '''
        set_cmd_text_color(BACKGROUND_WHITE|FOREGROUND_BLUE)
        sys.stdout.write(mess)
        resetColor()

# if __name__ == "__main__":
#     cmd_output = cmd_output()
#     cmd_output.printRed("你好\n")
